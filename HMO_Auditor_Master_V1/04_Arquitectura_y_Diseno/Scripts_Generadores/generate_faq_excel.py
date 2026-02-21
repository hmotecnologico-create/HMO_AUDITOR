import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
import os

def create_elite_faq_matrix():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz de Veracidad y FAQ"
    
    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Encabezados
    headers = ["Categoría", "Pregunta Clave", "Respuesta Técnica / Justificación", "Referencia Normativa"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    data = [
        ("Privacidad", "¿Cómo se garantiza que mis datos no salgan de la empresa?", "El sistema utiliza Ollama y ChromaDB operando 100% en local. No hay llamadas a APIs externas.", "ISO 27001 Cl. A.5.1"),
        ("Veracidad", "¿Por qué la IA no alucina al citar normas?", "Usamos RAG (Retrieval-Augmented Generation) anclado a PDFs oficiales. La IA cita el numeral exacto.", "ISO 19011:2018"),
        ("Persistencia", "¿Puedo retomar una auditoría de hace meses?", "Sí, cada empresa tiene un archivo 'audit_state.json' que guarda el progreso y la configuración.", "Continuidad de Negocio"),
        ("Integridad", "¿Cómo sé que el reporte no fue alterado?", "Cada documento generado se sella con un hash SHA-256 visible en el pie de página.", "Validez Jurídica"),
        ("Multi-Norma", "¿Cómo maneja varias normas a la vez?", "Usa 'Cross-Mapping' para identificar requisitos comunes y evitar duplicar la carga de evidencias.", "Eficiencia Operativa")
    ]

    for r_idx, row in enumerate(data, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Ajuste de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25

    output_path = r'd:\HMO\SENA\Auditor_Formatos\HMO_Auditor_Master_V1\06_Documentacion_Elite\Matriz_Veracidad_y_FAQ_Elite.xlsx'
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel FAQ Matrix generated at: {output_path}")

if __name__ == "__main__":
    create_elite_faq_matrix()
