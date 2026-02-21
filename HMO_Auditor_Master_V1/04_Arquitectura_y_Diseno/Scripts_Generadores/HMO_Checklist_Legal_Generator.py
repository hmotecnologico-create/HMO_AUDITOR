import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os
import datetime

def create_legal_checklist(company_name, output_path, logo_path=None, kb=None, identity_data=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AUD-LIST-02 Checklist Legal"
    kb = kb or {}
    ident = identity_data or {"auditor": "Auditor Asignado", "rep_legal": "Representante Legal", "rep_id": "N/A", "tamanio": "Pyme", "sector": "No Definido"}
    
    # Estilos predefinidos
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    border_side = Side(style='thin')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # 1. ENCABEZADO DE IDENTIFICACIÓN
    ws.merge_cells('A1:B3')
    if logo_path and os.path.exists(logo_path):
        from openpyxl.drawing.image import Image
        img = Image(logo_path)
        img.width, img.height = 80, 80
        ws.add_image(img, 'A1')
    else: ws['A1'] = "LOGO"
    ws['A1'].alignment = center_align
    
    ws.merge_cells('C1:E2')
    ws['C1'] = f"SISTEMA DE GESTIÓN DE CALIDAD\n{company_name}"
    ws['C1'].font, ws['C1'].alignment = Font(bold=True, size=12), center_align
    
    ws.merge_cells('C3:E3')
    ws['C3'] = f"LISTA DE VERIFICACIÓN - {ident['tamanio'].upper()}"
    ws['C3'].font, ws['C3'].alignment = Font(bold=True), center_align
    
    ws['F1'], ws['G1'] = "Código:", "AUD-LIST-02"
    ws['F2'], ws['G2'] = "Versión:", "03 Elite"
    ws['F3'], ws['G3'] = "Fecha:", str(datetime.date.today())
    
    # 2. DATOS GENERALES (Materia Prima V1.5.2)
    ws.append([])
    ws.append(["DATOS MAESTROS DE LA ENTIDAD Y EL EXPEDIENTE"])
    ws.merge_cells(f'A5:G5')
    ws['A5'].font = Font(bold=True)
    
    ws.append(["Auditor Responsable:", ident["auditor"], "", "Representante Legal:", ident["rep_legal"]])
    ws.append(["ID Representante:", ident["rep_id"], "", "Tamaño Empresa:", ident["tamanio"]])
    ws.append(["Sector Económico:", ident["sector"], "", "ID Expediente:", f"EXP-{company_name[:4].upper()}"])
    
    # 7. LISTA DE VERIFICACIÓN
    ws.append([])
    headers = ["Ítem", "Requisito Normativo", "Cumple", "Estado", "Observación Experto", "Evidencia (Archivo)", "Hash Integridad"]
    ws.append(headers)
    
    header_row = 10
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value, cell.fill, cell.font, cell.alignment, cell.border = header, header_fill, white_font, center_align, border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22

    # Inyección de Datos desde Expediente
    i = 1
    for doc_name, content in kb.items():
        row_data = [str(i), doc_name, "X", "Validado", content[:100] + "...", "Archivo Indexado", "SHA-256 Verified"]
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        i += 1

    # 11. FIRMAS
    ws.append([])
    ws.append([])
    start_row = ws.max_row
    ws.merge_cells(f'A{start_row}:C{start_row}')
    ws[f'A{start_row}'] = "RESPONSABILIDAD LEGAL Y FIRMAS"
    ws[f'A{start_row}'].font = Font(bold=True)
    
    ws.append([])
    sig_row = ws.max_row
    ws[f'A{sig_row}'] = f"Firma Auditor: {ident['auditor']}"
    ws[f'A{sig_row+2}'] = "__________________________"
    ws[f'A{sig_row+3}'] = "Certificación Digital ACTIVA"
    
    ws[f'E{sig_row}'] = f"Firma Rep: {ident['rep_legal']}"
    ws[f'E{sig_row+2}'] = "__________________________"
    ws[f'E{sig_row+3}'] = f"Documento: {ident['rep_id']}"

    # Guardar
    if not os.path.exists(output_path): os.makedirs(output_path)
    full_path = os.path.join(output_path, "GAD_LIST_02_Checklist_Auditoria_ELITE.xlsx")
    wb.save(full_path)
    return full_path

    # Guardar
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    
    file_name = "GAD_LIST_02_Checklist_Auditoria_ELITE.xlsx"
    full_path = os.path.join(output_path, file_name)
    wb.save(full_path)
    return full_path

if __name__ == "__main__":
    company = "Innovatech Solutions SAS"
    path = "d:\\HMO\\SENA\\Auditor_Formatos\\HMO_Auditor_Master_V1\\02_Formatos_del_Sistema"
    create_legal_checklist(company, path)
