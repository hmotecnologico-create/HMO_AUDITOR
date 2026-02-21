import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os

def create_audit_checklist(company_name, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GAD-LIST-02 Checklist"
    
    # --- Styles ---
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # --- Header Information ---
    ws.merge_cells('A1:E1')
    ws['A1'] = f"LISTA DE VERIFICACIÓN DE AUDITORÍA - {company_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = "Proceso:"
    ws['B2'] = "Desarrollo de Software"
    ws['D2'] = "Norma:"
    ws['E2'] = "ISO 9001:2015"
    
    # --- Columns Setup ---
    headers = ["ID", "Requisito Normativo (Bloqueado)", "Pregunta de Auditoría (Sugerencia AI)", "Cumplimiento (C/NC/NA)", "Evidencia y Hallazgos"]
    cols = ['A', 'B', 'C', 'D', 'E']
    
    for i, header in enumerate(headers):
        cell = ws.cell(row=4, column=i+1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border
        ws.column_dimensions[cols[i]].width = [10, 40, 40, 15, 50][i]

    # --- Sample Audit Content (AI Diligenced) ---
    data = [
        ["8.4.1", "La organización debe asegurarse de que los procesos suministrados externamente son conformes.", "¿Tiene la empresa un listado de proveedores de servicios en la nube evaluados?", "", ""],
        ["8.4.2", "La organización debe determinar los controles a aplicar a los proveedores externos.", "¿Se han definido criterios de aceptación para las librerías de terceros (Open Source)?", "", ""],
        ["8.5.1", "Control de la producción y de la provisión del servicio.", "¿Existe un registro de cambios (commit log) que valide la integridad del código fuente?", "", ""]
    ]
    
    for r_idx, row in enumerate(data, 5):
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=r_idx, column=c_idx+1)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
            
            # Protection Logic: Unlock only D and E columns
            if c_idx >= 3:
                cell.protection = openpyxl.styles.Protection(locked=False)
            else:
                cell.protection = openpyxl.styles.Protection(locked=True)

    # --- Sheet Protection ---
    ws.protection.sheet = True
    ws.protection.password = 'HMO_SAFE'
    
    # Final cleanup and save
    full_output = os.path.join(output_path, "GAD_LIST_02_Lista_Verificacion_Innovatech.xlsx")
    wb.save(full_output)
    print(f"Excel generado en: {full_output}")

if __name__ == "__main__":
    output_dir = r"d:\HMO\SENA\Auditor_Formatos\Formatos_Profesionales_HMO"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    create_audit_checklist("Innovatech Solutions SAS", output_dir)
